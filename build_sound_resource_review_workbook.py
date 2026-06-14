from collections import Counter
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


SOURCE = Path(r"C:\Users\rache\Downloads\sound_resource_guide_research.xlsx")
OUTPUT_DIR = Path("outputs/sound-resource-guide")
OUTPUT = OUTPUT_DIR / "sound_resource_guide_review_ready.xlsx"

MASTER_SHEET = "master_cleaned"

NEW_SHEETS = [
    "MVP_Approved_Seed_List",
    "Needs_Verification",
    "Top_List_Candidates",
    "Affiliate_Followup",
    "Low_Confidence_Review",
    "Launch_Summary",
]


def norm(value):
    return str(value or "").strip()


def yes(value):
    return norm(value).lower() == "yes"


def is_blank_or_unknown(value):
    return norm(value) == "" or norm(value).lower() == "unknown"


def as_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def read_records(ws):
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    records = []
    for excel_row, row in enumerate(rows[1:], start=2):
        record = dict(zip(headers, row))
        record["_excel_row"] = excel_row
        records.append(record)
    return headers, records


def sheet_records(ws, headers, records):
    ws.append(headers)
    for record in records:
        ws.append([record.get(header, "") for header in headers])


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        values = [norm(cell.value) for cell in column_cells[:200]]
        max_len = max([len(v) for v in values] + [10])
        width = min(max(max_len + 2, 12), 48)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def set_tab_color(ws, color):
    ws.sheet_properties.tabColor = color


def score_for_mvp(record):
    score = 0
    if norm(record.get("status")).lower() == "approved_candidate":
        score += 4
    if as_float(record.get("confidence_score")) >= 0.85:
        score += 3
    elif as_float(record.get("confidence_score")) >= 0.75:
        score += 1
    if yes(record.get("top_list_fit")):
        score += 2
    if yes(record.get("affiliate_program_available")):
        score += 2
    if norm(record.get("sponsor_potential")).lower() == "high":
        score += 2
    elif norm(record.get("sponsor_potential")).lower() == "medium":
        score += 1
    if is_blank_or_unknown(record.get("source_url")):
        score -= 5
    if as_float(record.get("confidence_score")) < 0.75:
        score -= 4
    return score


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(SOURCE)

    for sheet_name in NEW_SHEETS:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    master = wb[MASTER_SHEET]
    headers, records = read_records(master)

    approved = [
        r for r in records
        if norm(r.get("status")).lower() == "approved_candidate"
        and as_float(r.get("confidence_score")) >= 0.75
        and not is_blank_or_unknown(r.get("source_url"))
        and not is_blank_or_unknown(r.get("website_url"))
    ]
    approved_sorted = sorted(
        approved,
        key=lambda r: (score_for_mvp(r), as_float(r.get("confidence_score"))),
        reverse=True,
    )
    mvp_seed = approved_sorted[:150]

    needs_verification = [
        r for r in records
        if norm(r.get("status")).lower() == "needs_verification"
        or is_blank_or_unknown(r.get("source_url"))
        or is_blank_or_unknown(r.get("website_url"))
        or (
            yes(r.get("affiliate_program_available"))
            and is_blank_or_unknown(r.get("affiliate_signup_url"))
        )
        or as_float(r.get("confidence_score")) < 0.75
    ]

    top_list_candidates = [
        r for r in records
        if yes(r.get("top_list_fit"))
        and not is_blank_or_unknown(r.get("suggested_top_list"))
        and as_float(r.get("confidence_score")) >= 0.75
    ]
    top_list_candidates = sorted(
        top_list_candidates,
        key=lambda r: (norm(r.get("suggested_top_list")).lower(), -score_for_mvp(r)),
    )

    affiliate_followup = [
        r for r in records
        if yes(r.get("affiliate_program_available"))
        or norm(r.get("affiliate_program_available")).lower() == "unknown"
    ]
    affiliate_followup = sorted(
        affiliate_followup,
        key=lambda r: (
            0 if yes(r.get("affiliate_program_available")) else 1,
            is_blank_or_unknown(r.get("affiliate_signup_url")),
            norm(r.get("name")).lower(),
        ),
    )

    low_confidence = [
        r for r in records
        if as_float(r.get("confidence_score")) < 0.75
    ]
    low_confidence = sorted(low_confidence, key=lambda r: as_float(r.get("confidence_score")))

    tabs = [
        ("MVP_Approved_Seed_List", mvp_seed, "70AD47"),
        ("Needs_Verification", needs_verification, "FFC000"),
        ("Top_List_Candidates", top_list_candidates, "5B9BD5"),
        ("Affiliate_Followup", affiliate_followup, "A64D79"),
        ("Low_Confidence_Review", low_confidence, "C00000"),
    ]

    for sheet_name, sheet_data, tab_color in tabs:
        ws = wb.create_sheet(sheet_name)
        sheet_records(ws, headers, sheet_data)
        style_sheet(ws)
        set_tab_color(ws, tab_color)

    summary = wb.create_sheet("Launch_Summary", 0)
    set_tab_color(summary, "7030A0")
    summary_rows = [
        ["Sound Resource Guide Review Workbook", ""],
        ["Source workbook", str(SOURCE)],
        ["Total master records", len(records)],
        ["MVP approved seed rows", len(mvp_seed)],
        ["Needs verification rows", len(needs_verification)],
        ["Top-list candidate rows", len(top_list_candidates)],
        ["Affiliate follow-up rows", len(affiliate_followup)],
        ["Low-confidence rows", len(low_confidence)],
        ["", ""],
        ["Recommended next step", "Review MVP_Approved_Seed_List for taste, then verify Affiliate_Followup before publishing affiliate claims."],
        ["Quality note", "No individual people should be published; people/services entries should be platforms or websites only."],
        ["", ""],
        ["Top categories in MVP seed", "Count"],
    ]
    for category, count in Counter(norm(r.get("category")) for r in mvp_seed).most_common(15):
        summary_rows.append([category, count])

    for row in summary_rows:
        summary.append(row)
    summary["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="7030A0")
    summary["B1"].fill = PatternFill("solid", fgColor="7030A0")
    for cell in summary[13]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 90
    for row in summary.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUTPUT)
    print(OUTPUT.resolve())


if __name__ == "__main__":
    main()
