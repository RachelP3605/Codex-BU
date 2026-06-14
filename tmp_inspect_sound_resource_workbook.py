from openpyxl import load_workbook
from collections import Counter, defaultdict

path = r"C:\Users\rache\Downloads\sound_resource_guide_research.xlsx"
wb = load_workbook(path, read_only=True, data_only=True)

print("sheets:", wb.sheetnames)
for ws in wb.worksheets:
    print(f"--- {ws.title} ---")
    rows = list(ws.iter_rows(values_only=True))
    row_count = len(rows)
    col_count = max((len(row) for row in rows), default=0)
    print("rows", row_count, "cols", col_count)
    headers = list(rows[0]) if rows else []
    print("headers:", headers)
    for idx, row in enumerate(rows[1:4], start=2):
        print("row", idx, list(row[:8]))

print("--- QUALITY AUDIT: master_cleaned ---")
ws = wb["master_cleaned"]
rows = list(ws.iter_rows(values_only=True))
headers = list(rows[0])
records = [dict(zip(headers, row)) for row in rows[1:]]

def norm(value):
    return str(value or "").strip()

def is_blank(value):
    return norm(value) == "" or norm(value).lower() == "unknown"

print("total_records", len(records))
print("categories", Counter(norm(r.get("category")) for r in records).most_common(20))
print("status_counts", Counter(norm(r.get("status")) for r in records).most_common())
print("top_list_fit_counts", Counter(norm(r.get("top_list_fit")) for r in records).most_common())
print("affiliate_available_counts", Counter(norm(r.get("affiliate_program_available")) for r in records).most_common())

missing_source = [r for r in records if is_blank(r.get("source_url"))]
missing_website = [r for r in records if is_blank(r.get("website_url"))]
affiliate_yes_missing_signup = [
    r for r in records
    if norm(r.get("affiliate_program_available")).lower() == "yes"
    and is_blank(r.get("affiliate_signup_url"))
]
low_confidence = [
    r for r in records
    if float(r.get("confidence_score") or 0) < 0.75
]
duplicates = defaultdict(list)
for index, r in enumerate(records, start=2):
    key = norm(r.get("website_url")).lower().rstrip("/")
    if key:
        duplicates[key].append((index, norm(r.get("name"))))
duplicate_groups = {k: v for k, v in duplicates.items() if len(v) > 1}

print("missing_source_url", len(missing_source))
print("missing_website_url", len(missing_website))
print("affiliate_yes_missing_signup_url", len(affiliate_yes_missing_signup))
print("low_confidence_below_0_75", len(low_confidence))
print("duplicate_website_groups", len(duplicate_groups))

for label, items in [
    ("sample_missing_source", missing_source[:10]),
    ("sample_affiliate_yes_missing_signup", affiliate_yes_missing_signup[:10]),
    ("sample_low_confidence", low_confidence[:10]),
]:
    print(label, [norm(r.get("name")) for r in items])

print("sample_duplicate_groups")
for url, items in list(duplicate_groups.items())[:15]:
    print(url, items)
