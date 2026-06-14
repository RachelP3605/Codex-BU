from openpyxl import load_workbook

path = r"outputs\sound-resource-guide\sound_resource_guide_review_ready.xlsx"
wb = load_workbook(path, read_only=True, data_only=True)
ws = wb["agent_7_competitors_traffic"]
rows = list(ws.iter_rows(values_only=True))
headers = list(rows[0])

print(headers)
for row in rows[1:]:
    record = dict(zip(headers, row))
    if record.get("section") == "traffic_opportunity":
        print(record)
