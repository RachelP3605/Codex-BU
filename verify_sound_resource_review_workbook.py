from openpyxl import load_workbook

path = r"outputs\sound-resource-guide\sound_resource_guide_review_ready.xlsx"
wb = load_workbook(path, read_only=True, data_only=True)

print("sheets:", wb.sheetnames)
for sheet_name in [
    "Launch_Summary",
    "MVP_Approved_Seed_List",
    "Needs_Verification",
    "Top_List_Candidates",
    "Affiliate_Followup",
    "Low_Confidence_Review",
]:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    print(sheet_name, "data_rows", max(len(rows) - 1, 0), "columns", len(rows[0]) if rows else 0)
    if sheet_name == "Launch_Summary":
        for row in rows[:10]:
            print(row)
